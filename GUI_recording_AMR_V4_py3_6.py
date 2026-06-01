import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import os
import uuid
import tempfile
 
EXCEL_PATH = r"C:\Users\laura\OneDrive\Desktop\Python\GUI_AMR\Mockup_DataBase.xlsx"
# the r is to treat \ as characters otherwise i had to put \\

USER_OPTIONS     = ["Laura", "Bowei", "Daniel"]
MATERIAL_OPTIONS = ["Blocks", "Empty", "Spheres"]
 
COL_AMRcode  = 1   # columns of the properties Database
COL_Cp       = 65
COL_Porosity = 52
COL_Mass     = 50
 
WINDOW_WIDTH  = 1100
WINDOW_HEIGHT = 720
 
 
def load_all_data(path_to_excel):
    path_to_excel = str(path_to_excel)
    original_wb   = openpyxl.load_workbook(path_to_excel, data_only=True)
    #data_only=True  If a cell contains a formula, the last calculated value of the formula is returned
    
    sheet8 = original_wb.worksheets[0] #cascades
    sheet5 = original_wb.worksheets[1] #properties
 
    temp_wb   = openpyxl.Workbook()
 
    ws8       = temp_wb.active
    ws8.title = sheet8.title
    for row in sheet8.iter_rows(values_only=True):
        ws8.append(row)
 
    ws5 = temp_wb.create_sheet(title=sheet5.title)
    for r in range(1, sheet5.max_row + 1):
        amr_code       = sheet5.cell(row=r, column=COL_AMRcode).value
        cp_value       = sheet5.cell(row=r, column=COL_Cp).value
        porosity_value = sheet5.cell(row=r, column=COL_Porosity).value
        mass_value     = sheet5.cell(row=r, column=COL_Mass).value
        ws5.append([amr_code, cp_value, porosity_value, mass_value])
 
    original_wb.close()
 
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_path = temp_file.name
    temp_file.close()
 
    temp_wb.save(temp_path)
    temp_wb.close()
 
    wb    = openpyxl.load_workbook(temp_path, data_only=True)
    ws8_r = wb.worksheets[0]
 
    cascade_data = []
    for row in ws8_r.iter_rows(min_row=10, max_col=3, values_only=True):
        clean = [str(c) if c is not None else "" for c in row]
        while len(clean) < 3:
            clean.append("")
        cascade_data.append(clean)
 
    ws5_r      = wb.worksheets[1]
    amr_lookup = {}
    for row in ws5_r.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        row = list(row) + [""] * (4 - len(row))
        amr, cp, poro, mass = row[:4]
        if amr is not None:
            amr_lookup[str(amr).strip()] = (
                str(cp)   if isinstance(cp,   (int, float)) else "",
                str(poro) if isinstance(poro, (int, float)) else "",
                str(mass) if isinstance(mass, (int, float)) else "",
            )
 
    wb.close()
    if os.path.exists(temp_path):
        os.remove(temp_path)
 
    return cascade_data, amr_lookup
 
temp_path = load_all_data(EXCEL_PATH)
print("File saved at:", temp_path)






def lookup_amr_properties(amr_string, amr_lookup):
    codes   = [c.strip() for c in amr_string.split(",")]
    cp_list, po_list, ma_list = [], [], []
    for code in codes:
        cp, poro, mass = amr_lookup.get(code, ("", "", ""))
        cp_list.append(cp)
        po_list.append(poro)
        ma_list.append(mass)
    return ",".join(cp_list), ",".join(po_list), ",".join(ma_list)
 
# so it dosent show the whole path but it truncates it 
def truncate_path(path, max_chars=40):
    if len(path) <= max_chars:
        return path
    return "..." + path[-(max_chars - 3):]
 
 
def open_gui():
    cascade_data = []
    amr_lookup   = {}
    result       = {}
 
    root = tk.Tk()
    root.title("LabVIEW Data Entry")
    root.resizable(False, False)
 
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    x  = (sw - WINDOW_WIDTH)  // 2
    y  = (sh - WINDOW_HEIGHT) // 2
    root.geometry(f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}+{x}+{y}")
 
    left_frame = tk.Frame(root, bd=1, relief="sunken", bg="#f5f5f5", width=480)
    left_frame.grid(row=0, column=0, sticky="nsew", padx=(10,4), pady=10)
    left_frame.grid_propagate(False)
 
    right_outer = tk.Frame(root)
    right_outer.grid(row=0, column=1, sticky="nsew", padx=(4,10), pady=10)
 
    right_canvas = tk.Canvas(right_outer, highlightthickness=0)
    right_scroll = ttk.Scrollbar(right_outer, orient="vertical",
                                 command=right_canvas.yview)
    right_canvas.configure(yscrollcommand=right_scroll.set)
    right_scroll.pack(side="right", fill="y")
    right_canvas.pack(side="left", fill="both", expand=True)
 
    right_frame = tk.Frame(right_canvas, padx=14, pady=10)
    canvas_win  = right_canvas.create_window((0,0), window=right_frame, anchor="nw")
 
    def on_frame_configure(e):
        right_canvas.configure(scrollregion=right_canvas.bbox("all"))
    def on_canvas_configure(e):
        right_canvas.itemconfig(canvas_win, width=e.width)
 
    right_frame.bind("<Configure>", on_frame_configure)
    right_canvas.bind("<Configure>", on_canvas_configure)
 
    def on_mousewheel(e):
        right_canvas.yview_scroll(int(-1*(e.delta/120)), "units")
    right_canvas.bind_all("<MouseWheel>", on_mousewheel)
 
    root.columnconfigure(0, weight=2)
    root.columnconfigure(1, weight=2)
    root.rowconfigure(0, weight=1)
 
    # ── LEFT: cascade table ───────────────────────────────────────────────
    tk.Label(left_frame, text="Cascades", font=("Segoe UI", 10, "bold"),
             bg="#f5f5f5").grid(row=0, column=0, sticky="w", padx=8, pady=(8,2))
 
    cols = ("Cascade", "AMR1", "AMR2")
    tree = ttk.Treeview(left_frame, columns=cols, show="headings", selectmode="browse")
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=140, anchor="w", stretch=True)
 
    sb1 = ttk.Scrollbar(left_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=sb1.set)
    tree.grid(row=1, column=0, sticky="nsew", padx=(8,0), pady=(0,4))
    sb1.grid(row=1, column=1, sticky="ns", padx=(0,4), pady=(0,4))
 
    left_frame.columnconfigure(0, weight=1)
    left_frame.rowconfigure(1, weight=1)
    tree.tag_configure("highlight", background="#fff176")
 
    # ── LEFT: AMR properties table ────────────────────────────────────────
    tk.Label(left_frame, text="AMR Properties", font=("Segoe UI", 10, "bold"),
             bg="#f5f5f5").grid(row=2, column=0, sticky="w", padx=8, pady=(8,2))
 
    amr_cols = ("Code", "Cp", "Porosity", "Mass")
    amr_tree = ttk.Treeview(left_frame, columns=amr_cols, show="headings")
    for col in amr_cols:
        amr_tree.heading(col, text=col)
        amr_tree.column(col, width=110, anchor="w", stretch=True)
 
    sb2 = ttk.Scrollbar(left_frame, orient="vertical", command=amr_tree.yview)
    amr_tree.configure(yscrollcommand=sb2.set)
    amr_tree.grid(row=3, column=0, sticky="nsew", padx=(8,0), pady=(0,8))
    sb2.grid(row=3, column=1, sticky="ns", padx=(0,4), pady=(0,8))
    left_frame.rowconfigure(3, weight=1)
 
    # ── RIGHT: controls ───────────────────────────────────────────────────
    pad = dict(padx=4, pady=3)
    r   = 0
 
    def sep():
        nonlocal r
        ttk.Separator(right_frame, orient="horizontal").grid(
            row=r, column=0, columnspan=4, sticky="ew", pady=5)
        r += 1
 
    tk.Label(right_frame, text="User",
             font=("Segoe UI", 10, "bold")).grid(row=r, column=0, sticky="w", **pad)
    user_var = tk.StringVar(value=USER_OPTIONS[0])
    ttk.Combobox(right_frame, textvariable=user_var,
                 values=USER_OPTIONS, state="readonly", width=26).grid(
        row=r, column=1, columnspan=3, sticky="ew", **pad)
    r += 1
 
    tk.Label(right_frame, text="Material",
             font=("Segoe UI", 10, "bold")).grid(row=r, column=0, sticky="w", **pad)
    material_var = tk.StringVar(value=MATERIAL_OPTIONS[0])
    ttk.Combobox(right_frame, textvariable=material_var,
                 values=MATERIAL_OPTIONS, state="readonly", width=26).grid(
        row=r, column=1, columnspan=3, sticky="ew", **pad)
    r += 1
 
    sep()
 
    upload_btn = tk.Button(right_frame, text="Upload Cascades",
                           font=("Segoe UI", 10), width=22,
                           relief="groove", bg="#d0e8ff")
    upload_btn.grid(row=r, column=0, columnspan=4, pady=4)
    r += 1
 
    upload_status = tk.Label(right_frame, text="", font=("Segoe UI", 9), fg="#555")
    upload_status.grid(row=r, column=0, columnspan=4)
    r += 1
 
    sep()
 
    # ── File / folder pickers ─────────────────────────────────────────────
    def make_picker(label_text, grid_row, pick_folder=False):
        tk.Label(right_frame, text=label_text,
                 font=("Segoe UI", 10, "bold")).grid(
            row=grid_row, column=0, sticky="w", **pad)
 
        path_var    = tk.StringVar(value="")
        display_var = tk.StringVar(value="")
 
        entry = ttk.Entry(right_frame, textvariable=display_var,
                          state="readonly", width=28)
        entry.grid(row=grid_row, column=1, sticky="ew", padx=(0,2), pady=3)
 
        def browse():
            if pick_folder:
                path = filedialog.askdirectory(title=f"Select folder: {label_text}")
            else:
                path = filedialog.askopenfilename(
                    title=f"Select file: {label_text}",
                    filetypes=[("All files", "*.*")])
            if path:
                path_var.set(path)
                display_var.set(truncate_path(path))
 
        def clear():
            path_var.set("")
            display_var.set("")
 
        tk.Button(right_frame, text="...", font=("Segoe UI", 10),
                  command=browse, relief="groove", width=3).grid(
            row=grid_row, column=2, padx=(0,2), pady=3)
        tk.Button(right_frame, text="X", font=("Segoe UI", 9),
                  command=clear, relief="groove", width=2, fg="red").grid(
            row=grid_row, column=3, padx=(0,4), pady=3)
 
        return path_var
 
    housing_path_var      = make_picker("Housing Gen:",     r);                    r += 1
    inlet_outlet_path_var = make_picker("Inlet-Outlet:",    r);                    r += 1
    dispenser_path_var    = make_picker("Water dispenser:", r);                    r += 1
    blocks_cage_path_var  = make_picker("Blocks/Gd cage:",  r);                    r += 1
    frames_path_var       = make_picker("Frames:",          r, pick_folder=True);  r += 1
 
    cb_frame_widget = tk.Frame(right_frame)
    cb_frame_widget.grid(row=r, column=0, columnspan=4, sticky="w", padx=4, pady=4)
    r += 1
 
    frames_cb_var = tk.BooleanVar(value=False)
    tape_cb_var   = tk.BooleanVar(value=False)
 
    def on_frames_checked():
        if frames_cb_var.get():
            tape_cb_var.set(False)
    def on_tape_checked():
        if tape_cb_var.get():
            frames_cb_var.set(False)
 
    tk.Checkbutton(cb_frame_widget, text="Frames", variable=frames_cb_var,
                   font=("Segoe UI", 10),
                   command=on_frames_checked).pack(side="left", padx=(0,16))
    tk.Checkbutton(cb_frame_widget, text="Tape", variable=tape_cb_var,
                   font=("Segoe UI", 10),
                   command=on_tape_checked).pack(side="left")
 
    sep()
 
    tk.Label(right_frame, text="Cascade number",
             font=("Segoe UI", 10)).grid(row=r, column=0, sticky="w", **pad)
    cascade_num_var = tk.StringVar()
    ttk.Entry(right_frame, textvariable=cascade_num_var, width=26).grid(
        row=r, column=1, columnspan=3, sticky="ew", **pad)
    r += 1
 
    info_labels = {}
 
    for name in ["Cascade", "AMR1", "AMR2"]:
        lbl = tk.Label(right_frame, text=f"{name}: —",
                       font=("Segoe UI", 9), fg="#333",
                       anchor="w", width=42, justify="left")
        lbl.grid(row=r, column=0, columnspan=4, sticky="w", padx=4, pady=1)
        info_labels[name] = lbl
        r += 1
 
    sep()
 
    tk.Label(right_frame, text="AMR1 parameters",
             font=("Segoe UI", 9, "bold"), fg="#333").grid(
        row=r, column=0, columnspan=4, sticky="w", padx=4, pady=(0,2))
    r += 1
 
    for name in ["cpAMR1", "porosityAMR1", "massAMR1"]:
        f = tk.Frame(right_frame)
        f.grid(row=r, column=0, columnspan=4, sticky="w", padx=4, pady=1)
        tk.Label(f, text=f"{name}: ", font=("Segoe UI", 9, "bold"),
                 fg="#333", width=14, anchor="w").pack(side="left")
        val = tk.Label(f, text="—", font=("Segoe UI", 9), fg="#333",
                       width=24, anchor="w")
        val.pack(side="left")
        info_labels[name] = val
        r += 1
 
    sep()
 
    tk.Label(right_frame, text="AMR2 parameters",
             font=("Segoe UI", 9, "bold"), fg="#333").grid(
        row=r, column=0, columnspan=4, sticky="w", padx=4, pady=(0,2))
    r += 1
 
    for name in ["cpAMR2", "porosityAMR2", "massAMR2"]:
        f = tk.Frame(right_frame)
        f.grid(row=r, column=0, columnspan=4, sticky="w", padx=4, pady=1)
        tk.Label(f, text=f"{name}: ", font=("Segoe UI", 9, "bold"),
                 fg="#333", width=14, anchor="w").pack(side="left")
        val = tk.Label(f, text="—", font=("Segoe UI", 9), fg="#333",
                       width=24, anchor="w")
        val.pack(side="left")
        info_labels[name] = val
        r += 1
 
    sep()
 
    tk.Label(right_frame, text="Spheres AMR1",
             font=("Segoe UI", 10)).grid(row=r, column=0, sticky="w", **pad)
    spheres_amr1_var = tk.StringVar()
    ttk.Entry(right_frame, textvariable=spheres_amr1_var, width=26).grid(
        row=r, column=1, columnspan=3, sticky="ew", **pad)
    r += 1
 
    tk.Label(right_frame, text="Spheres AMR2",
             font=("Segoe UI", 10)).grid(row=r, column=0, sticky="w", **pad)
    spheres_amr2_var = tk.StringVar()
    ttk.Entry(right_frame, textvariable=spheres_amr2_var, width=26).grid(
        row=r, column=1, columnspan=3, sticky="ew", **pad)
    r += 1
 
    sep()
 
    # ── helpers ───────────────────────────────────────────────────────────
    # found_cascade stores cascade info + all 6 AMR property values
    found_cascade = {
        "cascade": "", "amr1": "", "amr2": "",
        "cp_amr1": "", "porosity_amr1": "", "mass_amr1": "",
        "cp_amr2": "", "porosity_amr2": "", "mass_amr2": "",
    }
 
    def clear_info_labels():
        for name in ["Cascade", "AMR1", "AMR2"]:
            info_labels[name].config(text=f"{name}: —")
        for name in ["cpAMR1", "porosityAMR1", "massAMR1",
                     "cpAMR2", "porosityAMR2", "massAMR2"]:
            info_labels[name].config(text="—")
        found_cascade.update(
            cascade="", amr1="", amr2="",
            cp_amr1="", porosity_amr1="", mass_amr1="",
            cp_amr2="", porosity_amr2="", mass_amr2="",
        )
 
    def filter_amr_table(amr1, amr2):
        amr_tree.delete(*amr_tree.get_children())
        codes = set()
        for part in [amr1, amr2]:
            for c in part.split(","):
                c = c.strip()
                if c:
                    codes.add(c)
        for code in codes:
            if code in amr_lookup:
                cp, poro, mass = amr_lookup[code]
                amr_tree.insert("", "end", values=(code, cp, poro, mass))
 
    def restore_amr_table():
        amr_tree.delete(*amr_tree.get_children())
        for code, (cp, poro, mass) in amr_lookup.items():
            amr_tree.insert("", "end", values=(code, cp, poro, mass))
 
    def update_info_labels(cascade, amr1, amr2):
        info_labels["Cascade"].config(text=f"Cascade: {cascade}"[:42])
        info_labels["AMR1"].config(text=f"AMR1: {amr1}"[:42])
        info_labels["AMR2"].config(text=f"AMR2: {amr2}"[:42])
 
        if amr_lookup:
            cp1, po1, ma1 = lookup_amr_properties(amr1, amr_lookup)
            cp2, po2, ma2 = lookup_amr_properties(amr2, amr_lookup)
        else:
            cp1 = po1 = ma1 = cp2 = po2 = ma2 = ""
 
        info_labels["cpAMR1"].config(text=(cp1 or "—")[:24])
        info_labels["porosityAMR1"].config(text=(po1 or "—")[:24])
        info_labels["massAMR1"].config(text=(ma1 or "—")[:24])
        info_labels["cpAMR2"].config(text=(cp2 or "—")[:24])
        info_labels["porosityAMR2"].config(text=(po2 or "—")[:24])
        info_labels["massAMR2"].config(text=(ma2 or "—")[:24])
 
        # store full values for returning to LabVIEW
        found_cascade.update(
            cascade=cascade, amr1=amr1, amr2=amr2,
            cp_amr1=cp1, porosity_amr1=po1, mass_amr1=ma1,
            cp_amr2=cp2, porosity_amr2=po2, mass_amr2=ma2,
        )
 
        filter_amr_table(amr1, amr2)
 
    def on_upload():
        nonlocal cascade_data, amr_lookup
        try:
            cascade_data, amr_lookup = load_all_data(EXCEL_PATH)
            for item in tree.get_children():
                tree.delete(item)
            for row in cascade_data:
                tree.insert("", "end", values=(row[0], row[1], row[2]))
            restore_amr_table()
            upload_status.config(
                text=f"{len(cascade_data)} cascades, {len(amr_lookup)} AMRs loaded",
                fg="green")
        except Exception as e:
            messagebox.showerror("Error loading file", str(e))
            upload_status.config(text="Load failed", fg="red")
 
    upload_btn.config(command=on_upload)
 
    trace_id_holder = {"id": None}
 
    def on_search(*args):
        query = cascade_num_var.get().strip()
        for item in tree.get_children():
            tree.item(item, tags=())
        if not query:
            clear_info_labels()
            restore_amr_table()
            return
        for item in tree.get_children():
            vals = tree.item(item, "values")
            if vals and str(vals[0]).startswith(query):
                tree.item(item, tags=("highlight",))
                tree.see(item)
                update_info_labels(vals[0], vals[1], vals[2])
                return
        info_labels["Cascade"].config(text="Cascade: not found")
        for name in ("AMR1", "AMR2"):
            info_labels[name].config(text=f"{name}: —")
        for name in ("cpAMR1", "porosityAMR1", "massAMR1",
                     "cpAMR2", "porosityAMR2", "massAMR2"):
            info_labels[name].config(text="—")
        restore_amr_table()
 
    trace_id_holder["id"] = cascade_num_var.trace_add("write", on_search)
 
    def on_tree_click(event):
        selected = tree.focus()
        if not selected:
            return
        vals = tree.item(selected, "values")
        if not vals:
            return
        for item in tree.get_children():
            tree.item(item, tags=())
        tree.item(selected, tags=("highlight",))
        update_info_labels(vals[0], vals[1], vals[2])
        cascade_num_var.trace_remove("write", trace_id_holder["id"])
        cascade_num_var.set(vals[0])
        trace_id_holder["id"] = cascade_num_var.trace_add("write", on_search)
 
    tree.bind("<<TreeviewSelect>>", on_tree_click)
 
    # ── LOAD button ───────────────────────────────────────────────────────
    def on_load():
        material = material_var.get()
        user     = user_var.get()
 
        if material == "Blocks":
            if not found_cascade["cascade"]:
                messagebox.showwarning(
                    "Missing data",
                    "Please upload cascades and enter a valid cascade number.")
                return
        elif material == "Spheres":
            if not spheres_amr1_var.get().strip() or \
               not spheres_amr2_var.get().strip():
                messagebox.showwarning(
                    "Missing data",
                    "Please fill in both Spheres AMR1 and Spheres AMR2.")
                return
 
        result.update(
            user          = user,
            material      = material,
            cascade       = found_cascade["cascade"],
            amr1_blocks   = found_cascade["amr1"],
            amr2_blocks   = found_cascade["amr2"],
            cp_amr1       = found_cascade["cp_amr1"],
            porosity_amr1 = found_cascade["porosity_amr1"],
            mass_amr1     = found_cascade["mass_amr1"],
            cp_amr2       = found_cascade["cp_amr2"],
            porosity_amr2 = found_cascade["porosity_amr2"],
            mass_amr2     = found_cascade["mass_amr2"],
            spheres_amr1  = spheres_amr1_var.get().strip(),
            spheres_amr2  = spheres_amr2_var.get().strip(),
            housing       = housing_path_var.get(),
            inlet_outlet  = inlet_outlet_path_var.get(),
            dispenser     = dispenser_path_var.get(),
            blocks_cage   = blocks_cage_path_var.get(),
            frames_folder = frames_path_var.get(),
            cb_frames     = frames_cb_var.get(),
            cb_tape       = tape_cb_var.get(),
        )
 
        root.destroy()
 
    tk.Button(right_frame, text="LOAD",
              font=("Segoe UI", 10, "bold"),
              bg="#2e7d32", fg="white",
              activebackground="#1b5e20", activeforeground="white",
              width=12, height=1,
              relief="flat", cursor="hand2",
              command=on_load).grid(row=r, column=0, columnspan=4, pady=10)
 
    right_frame.columnconfigure(1, weight=1)
    root.mainloop()
    return result
 
 
def run_gui():
    data = open_gui()
 
    # Always returns exactly 20 elements in the same fixed order.
    # Empty strings for anything not filled in.
    return [
        data.get("user",          ""),   # 0
        data.get("material",      ""),   # 1
        data.get("cascade",       ""),   # 2
        data.get("amr1_blocks",   ""),   # 3
        data.get("amr2_blocks",   ""),   # 4
        data.get("cp_amr1",       ""),   # 5
        data.get("porosity_amr1", ""),   # 6
        data.get("mass_amr1",     ""),   # 7
        data.get("cp_amr2",       ""),   # 8
        data.get("porosity_amr2", ""),   # 9
        data.get("mass_amr2",     ""),   # 10
        data.get("spheres_amr1",  ""),   # 11
        data.get("spheres_amr2",  ""),   # 12
        data.get("housing",       ""),   # 13
        data.get("inlet_outlet",  ""),   # 14
        data.get("dispenser",     ""),   # 15
        data.get("blocks_cage",   ""),   # 16
        data.get("frames_folder", ""),   # 17
        str(data.get("cb_frames", False)),  # 18
        str(data.get("cb_tape",   False)),  # 19
    ]
 
 
if __name__ == "__main__":
    output = run_gui()
    print("Returned to LabVIEW:", output)
